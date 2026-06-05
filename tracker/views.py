import profile

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.shortcuts import render, redirect, get_object_or_404
from datetime import date
from decimal import Decimal
from django.contrib.auth.models import User

from io import BytesIO
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import WorkDay
from .forms import WorkDayForm, WorkPeriodFormSet
import csv
from django.http import HttpResponse

@login_required
def work_day_list(request):
    user = request.user

    if user.groups.filter(name="Manager").exists() or user.is_superuser:
        days = WorkDay.objects.all()
    else:
        days = WorkDay.objects.filter(user=user)

    return render(request, "tracker/work_day_list.html", {"days": days})

@login_required
def work_day_create(request):
    if request.method == "POST":
        form = WorkDayForm(request.POST)
        formset = WorkPeriodFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            work_day = form.save(commit=False)
            work_day.user = request.user
            work_day.save()

            periods = formset.save(commit=False)
            for period in periods:
                period.work_day = work_day
                period.save()

            return redirect("work_day_list")
    else:
        form = WorkDayForm()
        formset = WorkPeriodFormSet()

    return render(request, "tracker/work_day_form.html", {
        "form": form,
        "formset": formset,
    })

@login_required
def work_day_update(request, pk):
    work_day = get_object_or_404(WorkDay, pk=pk)
    
    if work_day.approved and not is_manager(request.user):
        return redirect("work_day_list")

    if work_day.user != request.user and not request.user.groups.filter(name="Manager").exists() and not request.user.is_superuser:
        return redirect("work_day_list")

    if request.method == "POST":
        form = WorkDayForm(request.POST, instance=work_day)
        formset = WorkPeriodFormSet(request.POST, instance=work_day)

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect("work_day_list")
    else:
        form = WorkDayForm(instance=work_day)
        formset = WorkPeriodFormSet(instance=work_day)

    return render(request, "tracker/work_day_form.html", {
        "form": form,
        "formset": formset,
    })

@login_required
def work_day_delete(request, pk):
    work_day = get_object_or_404(WorkDay, pk=pk)

    if work_day.approved and not is_manager(request.user):
        return redirect("work_day_list")

    if work_day.user != request.user and not request.user.groups.filter(name="Manager").exists() and not request.user.is_superuser:
        return redirect("work_day_list")

    if request.method == "POST":
        work_day.delete()
        return redirect("work_day_list")

    return render(request, "tracker/work_day_confirm_delete.html", {
        "work_day": work_day
    })  

@login_required
def monthly_summary(request):
    user = request.user

    today = date.today()

    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    #hourly_rate = Decimal("30.00")
    profile = getattr(user, "userprofile", None)

    if profile:
        hourly_rate = profile.hourly_rate
    else:
        hourly_rate = Decimal("30.00")
        


    if user.groups.filter(name="Manager").exists() or user.is_superuser:
        days = WorkDay.objects.filter(date__year=year, date__month=month)
    else:
        days = WorkDay.objects.filter(user=user, date__year=year, date__month=month)

    total_minutes = sum(day.total_minutes() for day in days)
    total_hours = Decimal(total_minutes) / Decimal(60)
    salary = total_hours * hourly_rate

    context = {
        "days": days,
        "year": year,
        "month": month,
        "total_minutes": total_minutes,
        "total_hours": round(total_hours, 2),
        "hourly_rate": hourly_rate,
        "salary": round(salary, 2),
    }

    return render(request, "tracker/monthly_summary.html", context)


def is_manager(user):
    return user.groups.filter(name="Manager").exists() or user.is_superuser

# @login_required
# def manager_dashboard(request):
#     if not is_manager(request.user):
#         return redirect("work_day_list")

#     today = date.today()
#     year = int(request.GET.get("year", today.year))
#     month = int(request.GET.get("month", today.month))

#     days = WorkDay.objects.filter(
#         date__year=year,
#         date__month=month
#     ).select_related("user").prefetch_related("periods")

#     total_minutes = sum(day.total_minutes() for day in days)

#     context = {
#         "days": days,
#         "year": year,
#         "month": month,
#         "total_minutes": total_minutes,
#         "total_hours": round(total_minutes / 60, 2),
#     }

#     return render(request, "tracker/manager_dashboard.html", context)

@login_required
def manager_dashboard(request):
    if not is_manager(request.user):
        return redirect("work_day_list")

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))
    employee_id = request.GET.get("employee")

    days = WorkDay.objects.filter(
        date__year=year,
        date__month=month
    ).select_related("user").prefetch_related("periods")

    if employee_id:
        days = days.filter(user_id=employee_id)

    employees = User.objects.filter(
        workday__isnull=False
    ).distinct().order_by("username")

    total_minutes = sum(day.total_minutes() for day in days)

    context = {
        "days": days,
        "employees": employees,
        "selected_employee": employee_id,
        "year": year,
        "month": month,
        "total_minutes": total_minutes,
        "total_hours": round(total_minutes / 60, 2),
    }

    return render(request, "tracker/manager_dashboard.html", context)

@login_required
def approve_work_day(request, pk):
    if not is_manager(request.user):
        return redirect("work_day_list")

    work_day = get_object_or_404(WorkDay, pk=pk)

    if request.method == "POST":
        work_day.approved = True
        work_day.save()

    return redirect("manager_dashboard")


@login_required
def unapprove_work_day(request, pk):
    if not is_manager(request.user):
        return redirect("work_day_list")

    work_day = get_object_or_404(WorkDay, pk=pk)

    if request.method == "POST":
        work_day.approved = False
        work_day.save()

    return redirect("manager_dashboard")


def get_hourly_rate(user):
    profile = getattr(user, "userprofile", None)

    if profile:
        return profile.hourly_rate

    return Decimal("30.00")

@login_required
def export_employee_report_csv(request):
    user = request.user

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    days = WorkDay.objects.filter(
        user=user,
        date__year=year,
        date__month=month
    ).prefetch_related("periods")

    hourly_rate = get_hourly_rate(user)
    total_minutes = sum(day.total_minutes() for day in days)
    total_hours = Decimal(total_minutes) / Decimal(60)
    salary = total_hours * hourly_rate

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="raport_{user.username}_{year}_{month}.csv"'

    response.write("\ufeff")

    writer = csv.writer(response, delimiter=";")

    writer.writerow(["Raport czasu pracy"])
    writer.writerow(["Pracownik", user.username])
    writer.writerow(["Miesiac", month])
    writer.writerow(["Rok", year])
    writer.writerow(["Stawka godzinowa", f"{hourly_rate} zl/h"])
    writer.writerow(["Lacznie godzin", f"{round(total_hours, 2)}"])
    writer.writerow(["Wynagrodzenie", f"{round(salary, 2)} zl"])
    writer.writerow([])

    writer.writerow(["Data", "Przedzialy", "Razem minut", "Razem godzin", "Status"])

    for day in days:
        periods = ", ".join(
            f"{p.start_time.strftime('%H:%M')}-{p.end_time.strftime('%H:%M')}"
            for p in day.periods.all()
        )

        day_minutes = day.total_minutes()
        day_hours = Decimal(day_minutes) / Decimal(60)

        writer.writerow([
            day.date,
            periods,
            day_minutes,
            round(day_hours, 2),
            "Zatwierdzone" if day.approved else "Oczekuje",
        ])

    return response

@login_required
def export_manager_report_csv(request):
    if not is_manager(request.user):
        return redirect("work_day_list")

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))
    employee_id = request.GET.get("employee")

    days = WorkDay.objects.filter(
        date__year=year,
        date__month=month
    ).select_related("user").prefetch_related("periods")

    if employee_id:
        days = days.filter(user_id=employee_id)

    total_minutes = sum(day.total_minutes() for day in days)
    total_hours = Decimal(total_minutes) / Decimal(60)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="raport_szefa_{year}_{month}.csv"'

    response.write("\ufeff")

    writer = csv.writer(response, delimiter=";")

    writer.writerow(["Raport czasu pracy - panel szefa"])
    writer.writerow(["Miesiac", month])
    writer.writerow(["Rok", year])
    writer.writerow(["Lacznie godzin", f"{round(total_hours, 2)}"])
    writer.writerow([])

    writer.writerow([
        "Pracownik",
        "Data",
        "Przedzialy",
        "Razem minut",
        "Razem godzin",
        "Status"
    ])

    for day in days:
        periods = ", ".join(
            f"{p.start_time.strftime('%H:%M')}-{p.end_time.strftime('%H:%M')}"
            for p in day.periods.all()
        )

        day_minutes = day.total_minutes()
        day_hours = Decimal(day_minutes) / Decimal(60)

        writer.writerow([
            day.user.username,
            day.date,
            periods,
            day_minutes,
            round(day_hours, 2),
            "Zatwierdzone" if day.approved else "Oczekuje",
        ])

    return response




#Exel:
def decimal_hours_from_minutes(minutes):
    return round(Decimal(minutes) / Decimal(60), 2)


def format_periods(periods):
    return " ".join(
        f"{p.start_time.strftime('%H:%M')}-{p.end_time.strftime('%H:%M')}"
        for p in periods
    )


def format_hours_minutes(minutes):
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours:02d}:{mins:02d}"


def style_report_sheet(ws):

    ws.sheet_view.showGridLines = False

    thin_gray = Side(style="thin", color="D9D9D9")
    medium_gray = Side(style="medium", color="808080")

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="1F4E78")
    summary_fill = PatternFill("solid", fgColor="E2F0D9")
    weekend_fill = PatternFill("solid", fgColor="F2F2F2")
    approved_fill = PatternFill("solid", fgColor="D9EAD3")
    pending_fill = PatternFill("solid", fgColor="FFF2CC")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin_gray)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=14)
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in [4, 9]:
        for cell in ws[row_idx]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(top=medium_gray, bottom=medium_gray)

    for row in range(10, ws.max_row + 1):
        day_type = ws.cell(row=row, column=3).value
        status = ws.cell(row=row, column=7).value

        if day_type == "Wolne":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = weekend_fill

        if status == "Zatwierdzone":
            ws.cell(row=row, column=7).fill = approved_fill
        elif status == "Oczekuje":
            ws.cell(row=row, column=7).fill = pending_fill

    for row in range(ws.max_row - 4, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = summary_fill

    column_widths = {
        "A": 14,
        "B": 16,
        "C": 12,
        "D": 36,
        "E": 16,
        "F": 16,
        "G": 16,
        "H": 26,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 45
    ws.freeze_panes = "A10"


@login_required
def export_employee_report_xlsx(request):
    user = request.user

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    days = WorkDay.objects.filter(
        user=user,
        date__year=year,
        date__month=month
    ).prefetch_related("periods").order_by("date")

    days_by_date = {day.date.day: day for day in days}

    hourly_rate = get_hourly_rate(user)
    total_minutes = sum(day.total_minutes() for day in days)
    total_hours = Decimal(total_minutes) / Decimal(60)
    salary = total_hours * hourly_rate

    wb = Workbook()
    ws = wb.active
    ws.title = f"Raport {month:02d}-{year}"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"RAPORT CZASU PRACY - {month:02d}/{year}"

    ws["A3"] = "Pracownik"
    ws["B3"] = user.username
    ws["D3"] = "Miesiąc"
    ws["E3"] = f"{month:02d}/{year}"
    ws["G3"] = "Wygenerowano"
    ws["H3"] = date.today().strftime("%Y-%m-%d")

    ws.append([])
    ws.append(["Podsumowanie", "", "", "", "", "", "", ""])
    ws.append(["Łącznie godzin", float(round(total_hours, 2)), "", "Łącznie czasu", format_hours_minutes(total_minutes), "", "", ""])
    ws.append(["Stawka godzinowa", float(hourly_rate), "zł/h", "Wynagrodzenie", float(round(salary, 2)), "zł", "", ""])
    ws.append([])

    ws.append([
        "Data",
        "Dzień tygodnia",
        "Typ dnia",
        "Godziny pracy",
        "Czas [hh:mm]",
        "Czas [h]",
        "Status",
        "Opis",
    ])

    polish_days = {
        0: "Poniedziałek",
        1: "Wtorek",
        2: "Środa",
        3: "Czwartek",
        4: "Piątek",
        5: "Sobota",
        6: "Niedziela",
    }

    _, last_day = monthrange(year, month)

    for day_number in range(1, last_day + 1):
        current = date(year, month, day_number)
        work_day = days_by_date.get(day_number)
        is_weekend = current.weekday() >= 5

        if work_day:
            periods = list(work_day.periods.all())
            minutes = work_day.total_minutes()
            periods_text = format_periods(periods)
            status = "Zatwierdzone" if work_day.approved else "Oczekuje"
            description = work_day.description
            day_type = "Praca"
        else:
            minutes = 0
            periods_text = ""
            status = ""
            description = ""
            day_type = "Wolne" if is_weekend else ""

        ws.append([
            current.strftime("%Y-%m-%d"),
            polish_days[current.weekday()],
            day_type,
            periods_text,
            format_hours_minutes(minutes) if minutes else "",
            float(decimal_hours_from_minutes(minutes)) if minutes else "",
            status,
            description,
        ])

    ws.append([])
    ws.append(["SUMA", "", "", "", format_hours_minutes(total_minutes), float(round(total_hours, 2)), "", ""])
    ws.append(["STAWKA", "", "", "", "", float(hourly_rate), "zł/h", ""])
    ws.append(["DO WYPŁATY", "", "", "", "", float(round(salary, 2)), "zł", ""])
    ws.append([])
    ws.append(["Podpis pracownika", "", "", "Podpis przełożonego", "", "", "", ""])

    style_report_sheet(ws)

    ws["B6"].number_format = "0.00"
    ws["B7"].number_format = "0.00"
    ws["E7"].number_format = "0.00"
    ws[f"F{ws.max_row-3}"].number_format = "0.00"
    ws[f"F{ws.max_row-2}"].number_format = "0.00"
    ws[f"F{ws.max_row-1}"].number_format = "0.00"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="raport_{user.username}_{year}_{month:02d}.xlsx"'
    )

    return response


@login_required
def export_manager_report_xlsx(request):
    if not is_manager(request.user):
        return redirect("work_day_list")

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))
    employee_id = request.GET.get("employee")

    days = WorkDay.objects.filter(
        date__year=year,
        date__month=month
    ).select_related("user").prefetch_related("periods").order_by("user__username", "date")

    if employee_id:
        days = days.filter(user_id=employee_id)

    total_minutes = sum(day.total_minutes() for day in days)
    total_hours = Decimal(total_minutes) / Decimal(60)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Panel szefa {month:02d}-{year}"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"RAPORT CZASU PRACY - PANEL SZEFA - {month:02d}/{year}"

    ws["A3"] = "Zakres"
    ws["B3"] = f"{month:02d}/{year}"
    ws["D3"] = "Wygenerowano"
    ws["E3"] = date.today().strftime("%Y-%m-%d")

    ws.append([])
    ws.append(["Podsumowanie", "", "", "", "", "", "", ""])
    ws.append(["Łącznie czasu", format_hours_minutes(total_minutes), "", "Łącznie godzin", float(round(total_hours, 2)), "", "", ""])
    ws.append([])

    ws.append([
        "Pracownik",
        "Data",
        "Dzień tygodnia",
        "Godziny pracy",
        "Czas [hh:mm]",
        "Czas [h]",
        "Status",
        "Opis",
    ])

    polish_days = {
        0: "Poniedziałek",
        1: "Wtorek",
        2: "Środa",
        3: "Czwartek",
        4: "Piątek",
        5: "Sobota",
        6: "Niedziela",
    }

    employee_totals = {}

    for day in days:
        minutes = day.total_minutes()
        employee_totals.setdefault(day.user.username, 0)
        employee_totals[day.user.username] += minutes

        ws.append([
            day.user.username,
            day.date.strftime("%Y-%m-%d"),
            polish_days[day.date.weekday()],
            format_periods(day.periods.all()),
            format_hours_minutes(minutes),
            float(decimal_hours_from_minutes(minutes)),
            "Zatwierdzone" if day.approved else "Oczekuje",
            day.description,
        ])

    ws.append([])
    ws.append(["PODSUMOWANIE PRACOWNIKÓW", "", "", "", "", "", "", ""])
    ws.append(["Pracownik", "Czas [hh:mm]", "Czas [h]", "", "", "", "", ""])

    for username, minutes in employee_totals.items():
        ws.append([
            username,
            format_hours_minutes(minutes),
            float(decimal_hours_from_minutes(minutes)),
            "", "", "", "", ""
        ])

    ws.append([])
    ws.append(["SUMA", format_hours_minutes(total_minutes), float(round(total_hours, 2)), "", "", "", "", ""])

    style_report_sheet(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="raport_szefa_{year}_{month:02d}.xlsx"'
    )

    return response


